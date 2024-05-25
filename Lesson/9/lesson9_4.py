"""Прочитать сохраненный csv-файл и сохранить данные в excel-файл, 
кроме возраста - столбец с этими данными не нужен."""

import csv
from  openpyxl import Workbook

with open('users.csv', 'r') as read_csv_users_file:
    users_csv_file = csv.DictReader(read_csv_users_file)
    users_data = list(users_csv_file)


users_book = Workbook()
users_data_sheet = users_book.active

for col, person in enumerate(users_data, start = 2):
    users_data_sheet.cell(row = 1, column = col, value = f"Person {col-1}")

users_data_sheet['A2'] = 'ID'
users_data_sheet['A3'] = 'Name'
users_data_sheet['A4'] = 'Phone number'

for col_num in range(1, len(users_data) + 1):
    users_data_sheet.cell(row = 2, column = col_num + 1, value = users_data[col_num - 1]['ID'])
    users_data_sheet.cell(row = 3, column = col_num + 1, value = users_data[col_num - 1]['Name'])
    users_data_sheet.cell(row = 4, column = col_num + 1, value = users_data[col_num - 1]['Phone number'])

users_book.save("users_data.xlsx")